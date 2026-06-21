"""将补充用例写入 Excel"""
import openpyxl
from copy import copy

src = r'D:\网申助手\Swag_Labs_测试用例_设计阶段版.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb.active

# ===== 新增用例 =====
new_cases = [
    # --- 登录模块补充 ---
    ['TC-LOGIN-017', '登录模块', 'problem_user登录验证', '1. 用户访问登录页', '1. 输入用户名：problem_user\n2. 输入密码：secret_sauce\n3. 点击登录', '登录成功，但商品图片可能显示异常（已知Bug）'],
    ['TC-LOGIN-018', '登录模块', 'performance_glitch_user登录', '1. 用户访问登录页', '1. 输入用户名：performance_glitch_user\n2. 输入密码：secret_sauce\n3. 点击登录', '登录成功，但页面加载存在明显延迟（性能Bug）'],
    ['TC-LOGIN-019', '登录模块', 'error_user登录验证', '1. 用户访问登录页', '1. 输入用户名：error_user\n2. 输入密码：secret_sauce\n3. 点击登录', '登录成功，但部分操作可能报错（已知Bug）'],
    ['TC-LOGIN-020', '登录模块', 'visual_user登录验证', '1. 用户访问登录页', '1. 输入用户名：visual_user\n2. 输入密码：secret_sauce\n3. 点击登录', '登录成功，但部分元素样式可能异常（视觉Bug）'],
    ['TC-LOGIN-021', '登录模块', '密码最大长度边界测试', '1. 用户访问登录页', '1. 输入用户名：standard_user\n2. 输入密码：256位字符\n3. 点击登录', '系统合理处理（提示错误或限制输入）'],
    ['TC-LOGIN-022', '登录模块', 'XSS注入攻击测试', '1. 用户访问登录页', '1. 输入用户名：<script>alert(1)</script>\n2. 输入密码：任意值\n3. 点击登录', '系统不执行脚本，合理处理特殊字符'],
    ['TC-LOGIN-023', '登录模块', '登录后浏览器后退验证', '1. 用户已成功登录', '1. 登录成功后点击浏览器后退按钮', '停留在商品页或跳转登录页（不应出现异常页面）'],
    ['TC-LOGIN-024', '登录模块', '登录后关闭Tab重新打开', '1. 用户已成功登录', '1. 关闭浏览器Tab\n2. 新Tab打开 https://www.saucedemo.com/inventory.html', '根据Session状态跳转（通常需重新登录）'],

    # --- 登出模块（全新模块！） ---
    ['TC-LOGOUT-001', '登出模块', '侧边栏Logout正常登出', '1. 用户已登录', '1. 点击左上角汉堡菜单\n2. 点击Logout', '跳转回登录页，URL为 https://www.saucedemo.com/'],
    ['TC-LOGOUT-002', '登出模块', '登出后浏览器后退验证', '1. 用户已登出', '1. 登出后点击浏览器后退按钮', '不应回到已登录状态的商品页，应跳转登录页'],
    ['TC-LOGOUT-003', '登出模块', '登出后URL直接访问inventory', '1. 用户已登出', '1. 地址栏直接输入 /inventory.html 访问', '跳转回登录页或提示需要登录'],

    # --- 侧边栏菜单模块（全新模块！） ---
    ['TC-MENU-001', '侧边栏菜单', 'All Items菜单项功能', '1. 用户已登录，在商品详情页', '1. 点击汉堡菜单\n2. 点击All Items', '跳转回商品列表页 /inventory.html'],
    ['TC-MENU-002', '侧边栏菜单', 'About菜单项功能', '1. 用户已登录', '1. 点击汉堡菜单\n2. 点击About', '跳转至Sauce Labs官网信息页'],
    ['TC-MENU-003', '侧边栏菜单', 'Reset App State功能', '1. 用户已登录，购物车中有商品', '1. 点击汉堡菜单\n2. 点击Reset App State', '购物车清空，所有Add to cart按钮恢复初始状态'],
    ['TC-MENU-004', '侧边栏菜单', '菜单关闭功能', '1. 用户已登录，侧边栏打开', '1. 点击侧边栏外的X按钮或页面区域', '侧边栏关闭，回到当前页面'],

    # --- 商品列表补充 ---
    ['TC-LIST-011', '商品列表页', '列表页直接Add to cart', '1. 用户已登录', '1. 在商品列表页点击某商品的Add to cart按钮', '按钮变为Remove，购物车Badge数字+1'],
    ['TC-LIST-012', '商品列表页', '列表页直接Remove商品', '1. 用户已登录，某商品已在购物车', '1. 点击该商品的Remove按钮', '按钮恢复为Add to cart，购物车Badge数字-1'],
    ['TC-LIST-013', '商品列表页', '商品总数验证', '1. 用户已登录', '1. 打开商品列表页，统计商品数量', '共6个商品（Sauce Labs Backpack/Bike Light/Bolt T-Shirt/Fleece Jacket/Onesie/Red T-Shirt）'],

    # --- 商品详情补充 ---
    ['TC-DETAIL-006', '商品详情页', '多次快速点击Add to cart', '1. 用户已登录，商品未在购物车', '1. 快速连续点击Add to cart按钮3次', '购物车Badge数仅增加1（防重复提交）'],
    ['TC-DETAIL-007', '商品详情页', '商品描述文字完整显示', '1. 用户已登录', '1. 进入商品详情页，查看描述区', '描述文字完整，无截断或乱码'],

    # --- 购物车补充 ---
    ['TC-CART-007', '购物车', 'Badge计数跨页面一致性', '1. 用户已登录，购物车有2件商品', '1. 从购物车页→商品列表→商品详情→购物车', '页面跳转全程Badge始终显示2，无变化'],
    ['TC-CART-008', '购物车', '购物车总价计算验证', '1. 用户已登录，购物车有2件不同商品', '1. 进入购物车页\n2. 手动计算单价之和', '页面总价 = 各商品单价之和（不含税）'],

    # --- 结账补充 ---
    ['TC-CHECKOUT-007', '结账流程', 'LastName为空提交', '1. 购物车有商品', '1. FirstName: test, LastName: 空, ZipCode: 12345\n2. 点击Continue', '提示Error: Last Name is required'],
    ['TC-CHECKOUT-008', '结账流程', 'ZipCode格式非法', '1. 购物车有商品', '1. 输入非法ZipCode（如abcde、!@#）\n2. 点击Continue', '系统合理处理（接受或提示格式错误）'],

    # --- 安全测试补充 ---
    ['TC-SEC-001', '安全测试', '未登录直接访问inventory.html', '1. 用户未登录', '1. 浏览器直接访问 https://www.saucedemo.com/inventory.html', '跳转回登录页，不可绕过认证'],
    ['TC-SEC-002', '安全测试', '未登录直接访问cart.html', '1. 用户未登录', '1. 浏览器直接访问 https://www.saucedemo.com/cart.html', '跳转回登录页，不可绕过认证'],
    ['TC-SEC-003', '安全测试', '未登录直接访问checkout', '1. 用户未登录', '1. 浏览器直接访问 https://www.saucedemo.com/checkout-step-one.html', '跳转回登录页，不可绕过认证'],
    ['TC-SEC-004', '安全测试', '登录后切换用户Session测试', '1. 用户A已登录', '1. 复制当前URL到无痕窗口', '无痕窗口无法复用Session，需重新登录'],
]

# 获取最后一行
last_row = ws.max_row
next_row = last_row + 1

# Copy formatting from existing row (use R2 as template)
template_row = 2

# Write header style reference
for i, case in enumerate(new_cases):
    row = next_row + i
    for j, val in enumerate(case):
        cell = ws.cell(row=row, column=j+1)
        cell.value = val
        # Try to copy style from template row
        template_cell = ws.cell(row=template_row, column=j+1)
        if template_cell.has_style:
            cell.font = copy(template_cell.font)
            cell.alignment = copy(template_cell.alignment)
            cell.border = copy(template_cell.border)
            cell.fill = copy(template_cell.fill)

# Add a marker row with explanation
note_row = next_row + len(new_cases) + 1
ws.cell(row=note_row, column=1).value = '=== 补充说明 ==='
ws.cell(row=note_row+1, column=1).value = '以下用例在原始设计中需修正'
ws.cell(row=note_row+2, column=1).value = 'TC-LIST-002 分页：Swag Labs仅6件商品无分页 → 建议删除或改为"确认所有商品一页展示"'
ws.cell(row=note_row+3, column=1).value = 'TC-LIST-003/004/005 搜索：Swag Labs无搜索框 → 建议删除'
ws.cell(row=note_row+4, column=1).value = 'TC03-筛选-001~004 品类筛选：Swag Labs无品类筛选器 → 建议删除'
ws.cell(row=note_row+5, column=1).value = 'TC-COMPAT-003 Safari：Windows环境无法测试 → 建议改为Edge或标注未执行'

wb.save(src)
print(f'Done! Added {len(new_cases)} cases starting at row {next_row}')
print(f'Old total: {last_row - 1} cases')
print(f'New total: {next_row + len(new_cases) - 1} cases')
print(f'Note: Please review and remove the incorrect cases listed at the bottom of the sheet')
