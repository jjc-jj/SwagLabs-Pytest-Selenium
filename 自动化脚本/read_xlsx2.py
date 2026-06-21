import openpyxl, collections

wb = openpyxl.load_workbook(r'D:\网申助手\Swag_Labs_测试用例_设计阶段版.xlsx')
ws = wb.active

# Count by module
module_counts = collections.Counter()
all_rows = []
for r in range(2, ws.max_row + 1):
    tc_id = str(ws.cell(r, 1).value or '')
    module = str(ws.cell(r, 2).value or '')
    title = str(ws.cell(r, 3).value or '')

    # Skip note rows
    if tc_id.startswith('===') or tc_id.startswith('以下') or tc_id.startswith('TC-LIST-002') == False and tc_id.startswith('注意'):
        if '补充说明' in tc_id or '修正' in tc_id or '建议' in tc_id:
            continue

    if tc_id and not tc_id.startswith('===') and not tc_id.startswith('以下') and not tc_id.startswith('TC-LIST-002 分页') and not tc_id.startswith('TC03-筛选') and not tc_id.startswith('TC-COMPAT-003'):
        module_counts[module] += 1
        all_rows.append((tc_id, module, title))

print(f"Total valid cases: {len(all_rows)}")
print()
for mod, count in sorted(module_counts.items(), key=lambda x: -x[1]):
    print(f"  {mod}: {count}")

print()
print("=== All cases by module ===")
current_mod = None
for tc_id, mod, title in all_rows:
    if mod != current_mod:
        current_mod = mod
        print(f"\n--- {mod} ---")
    print(f"  {tc_id}: {title}")

# Also check for any remaining problematic items
print()
print("=== Quality Checks ===")

# Check for remaining non-existent features
problematic = []
for tc_id, mod, title in all_rows:
    lower = title.lower()
    if '分页' in title or '搜索' in title or '筛选' in title and tc_id.startswith('TC03-筛选'):
        problematic.append(f"  {tc_id}: {title} (might test non-existent feature)")
    if 'safari' in title.lower() and 'compat' in tc_id.lower():
        problematic.append(f"  {tc_id}: {title} (Safari on Windows)")

if problematic:
    print("POTENTIAL ISSUES:")
    for p in problematic:
        print(p)
else:
    print("No obvious issues found!")
