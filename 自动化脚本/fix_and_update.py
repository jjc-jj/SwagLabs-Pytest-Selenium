import openpyxl, collections

wb = openpyxl.load_workbook(r'D:\网申助手\Swag_Labs_测试用例_设计阶段版.xlsx')
ws = wb.active

# Step 1: Fix module names - merge all login cases under same name
# Check what the original login module name is (row 2)
original_login_module = ws.cell(2, 2).value
print(f"Original login module name (row 2): [{original_login_module}]")

# Find and fix all login cases with mismatched module name
fixed = 0
for r in range(2, ws.max_row + 1):
    tc_id = str(ws.cell(r, 1).value or '')
    module = str(ws.cell(r, 2).value or '')
    if tc_id.startswith('TC-LOGIN-') and '登录' in module:
        if module != original_login_module:
            ws.cell(r, 2).value = original_login_module
            fixed += 1

print(f"Fixed {fixed} login module name mismatches")

# Also check if TC-LIST cases are split
list_module_orig = None
for r in range(2, ws.max_row + 1):
    tc_id = str(ws.cell(r, 1).value or '')
    if tc_id == 'TC-LIST-001':
        list_module_orig = str(ws.cell(r, 2).value or '')
        break

if list_module_orig:
    fixed2 = 0
    for r in range(2, ws.max_row + 1):
        tc_id = str(ws.cell(r, 1).value or '')
        module = str(ws.cell(r, 2).value or '')
        if tc_id.startswith('TC-LIST-') and '列表' in module and module != list_module_orig:
            ws.cell(r, 2).value = list_module_orig
            fixed2 += 1
    print(f"Fixed {fixed2} list module name mismatches")

# Same for detail, cart, checkout
for prefix, keyword in [('TC-DETAIL-', '详情'), ('TC-CART-', '购物车'), ('TC-CHECKOUT-', '结账')]:
    orig_name = None
    for r in range(2, ws.max_row + 1):
        tc_id = str(ws.cell(r, 1).value or '')
        if tc_id.startswith(prefix) and keyword in str(ws.cell(r, 2).value or ''):
            orig_name = str(ws.cell(r, 2).value or '')
            break
    if orig_name:
        f = 0
        for r in range(2, ws.max_row + 1):
            tc_id = str(ws.cell(r, 1).value or '')
            module = str(ws.cell(r, 2).value or '')
            if tc_id.startswith(prefix) and keyword in module and module != orig_name:
                ws.cell(r, 2).value = orig_name
                f += 1
        if f > 0:
            print(f"Fixed {f} {prefix} module name mismatches")

wb.save(r'D:\网申助手\Swag_Labs_测试用例_设计阶段版.xlsx')
print()
print("=== Fixed Excel saved ===")

# Step 2: Recount
module_counts = collections.OrderedDict()
total = 0
for r in range(2, ws.max_row + 1):
    tc_id = str(ws.cell(r, 1).value or '')
    module = str(ws.cell(r, 2).value or '')
    if tc_id.startswith('TC-'):
        module_counts[module] = module_counts.get(module, 0) + 1
        total += 1

print(f"\nTotal valid cases: {total}")
print()
for mod, cnt in module_counts.items():
    print(f"  {mod}: {cnt}")
