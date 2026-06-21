"""Reorganize Excel by module - using TC-ID prefix for reliable matching"""
import openpyxl

src = r'D:\网申助手\Swag_Labs_测试用例_设计阶段版.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb.active

# Step 1: Read all valid cases
all_cases = []
for r in range(2, ws.max_row + 1):
    tc_id = str(ws.cell(r, 1).value or '').strip()
    if not tc_id.startswith('TC'):
        continue
    row_data = [str(ws.cell(r, c).value or '').strip() for c in range(1, 7)]
    all_cases.append(row_data)

print(f"Read {len(all_cases)} valid cases")

# Step 2: Map TC-ID prefix to canonical module name
def get_module(tc_id):
    t = tc_id.upper()
    if 'LOGIN' in t: return ('01_登录模块', '登录模块')
    if 'LOGOUT' in t: return ('07_登出模块', '登出模块')
    if 'MENU' in t: return ('08_侧边栏菜单', '侧边栏菜单')
    if 'SEC-' in t: return ('09_安全测试', '安全测试')
    if 'LIST' in t: return ('02_商品列表页', '商品列表页')
    if 'SORT' in t or ('排序' in tc_id): return ('03_商品排序', '商品排序')
    if 'DETAIL' in t: return ('04_商品详情页', '商品详情页')
    if 'CART' in t: return ('05_购物车', '购物车')
    if 'CHECKOUT' in t or ('结账' in tc_id): return ('06_结账流程', '结账流程')
    if 'COMPAT' in t: return ('10_兼容性测试', '兼容性测试')
    if '边界' in tc_id: return ('11_边界值专项', '边界值专项')
    return ('99_其他', '其他')

# Apply and fix module names
for case in all_cases:
    sort_key, display_name = get_module(case[0])
    case.append(sort_key)  # column 7 for sorting
    case[1] = display_name  # Fix module name

# Step 3: Sort
all_cases.sort(key=lambda c: (c[6], c[0]))

# Step 4: Write back
# Delete old rows
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

for i, case in enumerate(all_cases):
    row = 2 + i
    for j in range(6):  # Only write first 6 columns
        ws.cell(row=row, column=j+1, value=case[j])

# Summary
summary_start = 2 + len(all_cases) + 1
ws.cell(row=summary_start, column=1, value='=== 模块统计 ===')
r = summary_start + 1

# Count
from collections import OrderedDict
mod_counts = OrderedDict()
for c in all_cases:
    mod = c[1]
    mod_counts[mod] = mod_counts.get(mod, 0) + 1

for mod, cnt in mod_counts.items():
    cases_in_mod = [c[0] for c in all_cases if c[1] == mod]
    ws.cell(row=r, column=1, value=f'{mod}: {cnt}条')
    ws.cell(row=r, column=2, value=f'{cases_in_mod[0]} ~ {cases_in_mod[-1]}')
    r += 1

ws.cell(row=r+1, column=1, value=f'合计: {len(all_cases)}条 | {len(mod_counts)}个模块')

wb.save(src)

print(f"\nSorted {len(all_cases)} cases into {len(mod_counts)} modules:")
for mod, cnt in mod_counts.items():
    cases_in_mod = [c[0] for c in all_cases if c[1] == mod]
    print(f"  {mod}: {cnt}条 ({cases_in_mod[0]} ~ {cases_in_mod[-1]})")
