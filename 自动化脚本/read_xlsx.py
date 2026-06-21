import openpyxl
wb = openpyxl.load_workbook(r'D:\网申助手\Swag_Labs_测试用例_设计阶段版.xlsx')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'=== Sheet: {name} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    print(f'Headers: {headers}')
    print()
    for r in range(2, ws.max_row+1):
        row_data = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column+1)]
        print(f'  R{r}: ' + ' | '.join(row_data))
    print()
