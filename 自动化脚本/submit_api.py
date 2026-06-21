"""Submit 11 defects to ZenTao via REST API (DEF-003~DEF-013)"""
import requests, urllib3, time, json, os
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

urllib3.disable_warnings()
BASE = "https://zenboard.demo.qucheng.cc"
API_URL = f"{BASE}/api.php/v1/bugs"
EXCEL = "Swag_Labs_测试资产_企业级规范版.xlsx"

SEV_MAP = {'致命':1, '严重':2, '一般':3, '轻微':4}
PRI_MAP = {'P0':1, 'P1':2, 'P2':3, 'P3':4}
TYPE_MAP = {'功能性缺陷':'codeerror','视觉缺陷':'design','安全隐患':'security',
            'UX缺陷':'others','校验缺陷':'codeerror','展示缺陷':'design','功能缺陷':'codeerror'}

# Read defects from Excel
wb = openpyxl.load_workbook(EXCEL)
ws = wb[wb.sheetnames[1]]
defects = []
for row in ws.iter_rows(min_row=4, max_row=16, values_only=True):
    if row[0] and str(row[0]).startswith('DEF-'):
        defects.append({
            'defect_id':str(row[0]),'tc_id':str(row[1] or ''),'title':str(row[2] or ''),
            'severity':str(row[4] or '一般'),'priority':str(row[5] or 'P2'),
            'type':str(row[6] or '功能性缺陷'),'env':str(row[7] or ''),
            'precondition':str(row[8] or ''),'steps':str(row[9] or ''),
            'expected':str(row[10] or ''),'actual':str(row[11] or ''),
            'impact':str(row[12] or ''),'fix':str(row[13] or ''),
        })
wb.close()

# Filter: skip DEF-001, DEF-002 (already manually submitted)
targets = [d for d in defects if d['defect_id'] not in ['DEF-001','DEF-002']]
print(f"Loaded {len(defects)} defects, submitting {len(targets)} (DEF-003 ~ DEF-013)")

# Login via Selenium
print("\nAuthenticating...")
opts = webdriver.EdgeOptions()
opts.add_argument('--ignore-certificate-errors')
opts.add_experimental_option('excludeSwitches', ['enable-logging'])
drv = webdriver.Edge(options=opts)
drv.get(f"{BASE}/index.php?m=user&f=login")
time.sleep(2)
drv.find_element(By.ID, "submit").click()
time.sleep(3)
print(f"Logged in: {drv.title}")

s = requests.Session()
s.verify = False
for c in drv.get_cookies():
    s.cookies.set(c['name'], c['value'], domain=c.get('domain',''))
drv.quit()

def build_steps_html(d):
    """Build steps in ZenTao HTML format"""
    sections = [
        ('[前置条件]', d.get('precondition','')),
        ('[复现步骤]', d.get('steps','')),
        ('[预期结果]', d.get('expected','')),
        ('[实际结果]', d.get('actual','')),
        ('[影响范围]', d.get('impact','')),
        ('[建议修复方案]', d.get('fix','')),
    ]
    result = []
    for label, content in sections:
        if content and content.strip():
            result.append(f'<p><strong>{label}</strong></p>')
            for line in content.split('\n'):
                line = line.strip()
                if line:
                    result.append(f'<p>{line}</p>')
    result.append(f'<p><strong>[测试环境]</strong> {d.get("env","")}</p>')
    result.append(f'<p><strong>[关联用例]</strong> {d.get("tc_id","")}</p>')
    result.append(f'<p><strong>[缺陷ID]</strong> {d.get("defect_id","")}</p>')
    return '\n'.join(result)

# Submit each bug
ok, bad = 0, []
results = []

for i, d in enumerate(targets):
    n = d['defect_id']
    print(f"\n[{i+1}/11] {n}: {d['title'][:55]}...")

    payload = {
        'title': d['title'],
        'product': 8,
        'project': 24,
        'openedBuild': '主干',
        'severity': SEV_MAP.get(d['severity'], 3),
        'pri': PRI_MAP.get(d['priority'], 3),
        'type': TYPE_MAP.get(d['type'], 'codeerror'),
        'steps': build_steps_html(d),
        'os': 'Windows 11',
        'browser': 'Edge 149',
        'keywords': f'{d["defect_id"]},{d["tc_id"]}',
        'color': '#0000FF',
    }

    try:
        r = s.post(API_URL, json=payload, timeout=30)
        if r.status_code == 201:
            bug_id = r.json().get('id', '?')
            print(f"  CREATED (Bug #{bug_id})")
            ok += 1
            results.append({'defect': n, 'bug_id': bug_id, 'status': 'OK'})
        else:
            try:
                err = r.json().get('error', r.text[:150])
            except:
                err = r.text[:150]
            print(f"  FAILED ({r.status_code}): {err}")
            bad.append(n)
            results.append({'defect': n, 'bug_id': None, 'status': str(err)})
    except Exception as e:
        print(f"  ERROR: {e}")
        bad.append(n)
        results.append({'defect': n, 'bug_id': None, 'status': str(e)})

    time.sleep(0.5)

# Summary
print(f"\n{'='*60}")
print(f"SUBMISSION COMPLETE")
print(f"  Success: {ok}/{len(targets)}")
if bad:
    print(f"  Failed: {', '.join(bad)}")
print(f"  + DEF-001/002 already done = {ok+2}/13 total")
print(f"\nSubmitted bugs:")
for r in results:
    print(f"  {r['defect']}: Bug #{r['bug_id']} - {r['status']}")
print(f"\nBug list: {BASE}/index.php?m=project&f=bug&projectID=162")
